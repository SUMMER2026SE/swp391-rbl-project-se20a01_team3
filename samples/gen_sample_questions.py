# -*- coding: utf-8 -*-
"""
Sinh các file Excel câu hỏi mẫu đúng định dạng import của Bee Academy.

Định dạng cột (khớp với ExcelImportModal.tsx — hàng 1 là header, dữ liệu từ hàng 2):
    A: Nội dung câu hỏi   (bắt buộc)
    B: Loại               TN = trắc nghiệm | DS = đúng/sai
    C: Độ khó             D = dễ | TB = trung bình | K = khó
    D: Đáp án A           (bắt buộc với TN)
    E: Đáp án B           (bắt buộc với TN)
    F: Đáp án C           (tùy chọn)
    G: Đáp án D           (tùy chọn)
    H: Đáp án đúng        A / B / C / D  (với DS: A = Đúng, B = Sai)
    I: Giải thích         (tùy chọn)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

HEADER = [
    "Nội dung câu hỏi",
    "Loại (TN/DS)",
    "Độ khó (D/TB/K)",
    "Đáp án A",
    "Đáp án B",
    "Đáp án C",
    "Đáp án D",
    "Đáp án đúng (A/B/C/D)",
    "Giải thích (tùy chọn)",
]

COL_WIDTHS = [52, 14, 16, 22, 22, 22, 22, 22, 40]

# ── Chương 1: Phép nhân & phép chia đa thức ──────────────────────────────────
CHUONG_1 = [
    ["Kết quả của phép nhân đơn thức 3x²·2x³ là gì?", "TN", "D",
     "6x⁵", "5x⁶", "6x⁶", "5x⁵", "A",
     "Nhân hệ số 3·2=6, cộng số mũ 2+3=5 → 6x⁵."],
    ["Khai triển hằng đẳng thức (a + b)² bằng:", "TN", "D",
     "a² + b²", "a² + 2ab + b²", "a² - 2ab + b²", "2a² + 2b²", "B",
     "Bình phương của một tổng: (a+b)² = a² + 2ab + b²."],
    ["Khai triển (x - 3)² ta được:", "TN", "TB",
     "x² - 9", "x² - 6x + 9", "x² + 6x + 9", "x² - 3x + 9", "B",
     "(x-3)² = x² - 2·3·x + 9 = x² - 6x + 9."],
    ["Phân tích đa thức x² - 16 thành nhân tử:", "TN", "TB",
     "(x - 4)(x + 4)", "(x - 16)(x + 1)", "(x - 8)(x + 2)", "(x - 4)²", "A",
     "Hiệu hai bình phương: x² - 4² = (x-4)(x+4)."],
    ["Hằng đẳng thức (a - b)(a + b) bằng a² + b².", "DS", "D",
     "", "", "", "", "B",
     "Sai. Đúng phải là (a-b)(a+b) = a² - b²."],
    ["Đa thức 2x³ - 4x² + 6x chia hết cho đơn thức 2x.", "DS", "TB",
     "", "", "", "", "A",
     "Đúng. Mỗi hạng tử đều chứa nhân tử 2x → 2x(x² - 2x + 3)."],
    ["Giá trị của biểu thức x² - 2x + 1 tại x = 5 là:", "TN", "TB",
     "16", "9", "25", "36", "A",
     "x² - 2x + 1 = (x-1)² = (5-1)² = 16."],
    ["Phân tích x² + 5x + 6 thành nhân tử:", "TN", "K",
     "(x + 2)(x + 3)", "(x + 1)(x + 6)", "(x - 2)(x - 3)", "(x + 5)(x + 1)", "A",
     "Tìm hai số có tổng 5, tích 6 → 2 và 3 → (x+2)(x+3)."],
]

# ── Chương 2: Phân thức đại số ───────────────────────────────────────────────
CHUONG_2 = [
    ["Phân thức nào sau đây xác định khi x ≠ 2?", "TN", "D",
     "1/(x-2)", "1/(x+2)", "x/2", "(x-2)/3", "A",
     "Mẫu x-2 ≠ 0 ⇔ x ≠ 2 nên 1/(x-2) là đáp án."],
    ["Rút gọn phân thức (2x)/(4x²) (với x ≠ 0):", "TN", "TB",
     "1/(2x)", "2/x", "x/2", "1/2", "A",
     "(2x)/(4x²) = 2x/(4x·x) = 1/(2x)."],
    ["Điều kiện xác định của phân thức (x+1)/(x²-9) là:", "TN", "K",
     "x ≠ 3", "x ≠ -3", "x ≠ ±3", "x ≠ 9", "C",
     "Mẫu x²-9 = (x-3)(x+3) ≠ 0 ⇔ x ≠ 3 và x ≠ -3."],
    ["Hai phân thức a/b và c/d bằng nhau khi a·d = b·c.", "DS", "TB",
     "", "", "", "", "A",
     "Đúng. Đây là định nghĩa hai phân thức bằng nhau."],
    ["Phân thức 5/0 là một phân thức hợp lệ.", "DS", "D",
     "", "", "", "", "B",
     "Sai. Mẫu thức phải khác 0, nên 5/0 không xác định."],
    ["Kết quả phép cộng 1/x + 2/x (với x ≠ 0) là:", "TN", "D",
     "3/x", "3/(2x)", "2/x²", "3/x²", "A",
     "Cùng mẫu nên cộng tử: (1+2)/x = 3/x."],
    ["Mẫu thức chung của 1/(x) và 1/(x+1) là:", "TN", "TB",
     "x", "x + 1", "x(x+1)", "2x+1", "C",
     "Hai mẫu khác nhau và không có nhân tử chung → MTC = x(x+1)."],
]

# ── Chương 3: Phương trình bậc nhất một ẩn ───────────────────────────────────
CHUONG_3 = [
    ["Nghiệm của phương trình 2x - 6 = 0 là:", "TN", "D",
     "x = 2", "x = 3", "x = -3", "x = 6", "B",
     "2x = 6 ⇒ x = 3."],
    ["Phương trình bậc nhất một ẩn có dạng tổng quát:", "TN", "D",
     "ax² + bx + c = 0", "ax + b = 0 (a ≠ 0)", "ax + b = 0 (a = 0)", "a/x = b", "B",
     "Phương trình bậc nhất một ẩn: ax + b = 0 với a ≠ 0."],
    ["Giải phương trình 3x + 5 = 2x + 9, nghiệm là:", "TN", "TB",
     "x = 4", "x = -4", "x = 2", "x = 14", "A",
     "3x - 2x = 9 - 5 ⇒ x = 4."],
    ["Phương trình 0x = 5 vô nghiệm.", "DS", "TB",
     "", "", "", "", "A",
     "Đúng. Không có giá trị x nào thỏa 0 = 5 nên vô nghiệm."],
    ["Phương trình x + 3 = x + 3 có vô số nghiệm.", "DS", "TB",
     "", "", "", "", "A",
     "Đúng. Hai vế luôn bằng nhau với mọi x → vô số nghiệm."],
    ["Điều kiện xác định của phương trình 1/(x-2) = 3 là:", "TN", "K",
     "x ≠ 0", "x ≠ 2", "x ≠ 3", "x ≠ -2", "B",
     "Mẫu x - 2 ≠ 0 ⇔ x ≠ 2."],
    ["Nghiệm của phương trình 5(x - 1) = 2x + 7 là:", "TN", "K",
     "x = 4", "x = 3", "x = 2", "x = -4", "A",
     "5x - 5 = 2x + 7 ⇒ 3x = 12 ⇒ x = 4."],
]


def build_sheet(ws, rows):
    header_fill = PatternFill("solid", fgColor="C0392B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="center")

    ws.append(HEADER)
    for ci, _ in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS[ci - 1]

    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = wrap
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def make_file(path, sheet_name, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    build_sheet(ws, rows)
    wb.save(path)
    print(f"  ✔ {os.path.basename(path)} — {len(rows)} câu")


def main():
    out = os.path.dirname(os.path.abspath(__file__))
    print("Đang sinh file câu hỏi mẫu vào:", out)

    make_file(os.path.join(out, "cauhoi_chuong1_phepnhanchia_dathuc.xlsx"),
              "Chương 1", CHUONG_1)
    make_file(os.path.join(out, "cauhoi_chuong2_phanthuc.xlsx"),
              "Chương 2", CHUONG_2)
    make_file(os.path.join(out, "cauhoi_chuong3_phuongtrinh.xlsx"),
              "Chương 3", CHUONG_3)

    # File gộp tất cả — tiện test import số lượng lớn 1 lần
    make_file(os.path.join(out, "cauhoi_tatca_toan8.xlsx"),
              "Tất cả", CHUONG_1 + CHUONG_2 + CHUONG_3)

    print("Hoàn tất.")


if __name__ == "__main__":
    main()
